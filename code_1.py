import openpyxl
import os
from python_terraform import *

workbook=openpyxl.load_workbook("example.xlsx")

tf = Terraform(working_dir="./")  
tf.init()   

for worksheet in workbook.worksheets:

    activesheet=workbook[worksheet.title]
    resource_values_from_client={}
    resource_values_from_client["resource"]=worksheet.title

    for row in range(1, activesheet.max_row + 1):
        key = activesheet.cell(row, 1).value
        value = activesheet.cell(row, 2).value
        resource_values_from_client[key] = value

    if resource_values_from_client.get('need')==1:
        filename = "client_val.tfvars"
        
        fileWithPath = os.path.join("./Modules/{}/".format(resource_values_from_client.get("resource")), filename)
        
        with open(fileWithPath, "w") as file:
            for key,value in resource_values_from_client.items():
                if key=="resource" or key=="need":
                    continue
                stuff_in_string = "{} = \"{}\" \n".format(key, value)
                print(stuff_in_string)
                file.write(stuff_in_string)  
                print("check")
        
        terra_working_dir="./Modules/{}/".format(resource_values_from_client.get("resource"))
        tf = Terraform(working_dir=terra_working_dir)  
        tf.init()
        print(tf.plan())    
    else:
        print("no need")            
    
        
        


        